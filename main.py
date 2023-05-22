from openpyxl import Workbook
from openpyxl import load_workbook
import pickle
import shutil
from datetime import datetime
import pyautogui as pyag
import tkinter as tk
import os
from string import ascii_uppercase
from tkinter.filedialog import askopenfilename


# data structure
# f = [chunk, ...]
# chunk = {'name': name,
# 'data': {datetime: row[1:9], ...}
# }

def addlist(a, b):
    if len(a) != len(b):
        raise Exception(f'unequal lengths of \n{a}\n{b}')
    r = []
    for i in range(len(a)):
        r.append(a[i] + b[i])
    return r


class Compiler:
    def __init__(self):
        self.wb = Workbook()
        self.finals = []
        self.load_data()

    def load_data(self, file='data.pickle'):
        self.alldata = pickle.load(open(file, 'rb'))

    def compile1(self):
        tocompile = []
        for r, d, fs in os.walk('./tocompile'):
            for f in fs:
                ext = list(reversed(list(f.split('.'))))[0]
                if ext == 'xlsx':
                    tocompile.append(f)
        for f in tocompile:
            sheet = load_workbook(f'./tocompile/{f}').active
            n = 0
            while not sheet[f'A{n + 3}'].value in [None, ''] or str(sheet[f'A{n + 3}'].value).replace(' ', '') == '':
                n += 1
            final = {'name': sheet['B1'].value, 'data': {}}
            for i in range(n):
                row = []
                for letter in ascii_uppercase[0:9]:
                    value = sheet[f'{letter}{i + 3}'].value
                    try:
                        if letter != 'A':
                            float(value)
                        else:
                            row.append(value)
                            continue
                    except:
                        if value not in [None, '-', 'N/A']:
                            print(f'{f}: {f"{letter}{i + 3}"} = {value}')
                        value = 0
                    row.append(float(value))
                try:  # layer 2
                    final['data'][row[0]] = addlist(final['data'][row[0]], row[1:])
                except:
                    final['data'][row[0]] = row[1:]
            self.finals.append(final)  # layer 1
        pickle.dump(self.finals, open('data.pickle', 'wb'))
        # pickle.dump([], open('data.pickle', 'wb'))

    def compile2(self):
        tocompile = []
        self.load_data()
        for r, d, fs in os.walk('./thisweek'):
            for f in fs:
                ext = list(reversed(list(f.split('.'))))[0]
                if ext == 'xlsx':
                    tocompile.append(f)

        for f in tocompile:
            sheet = load_workbook(f'./thisweek/{f}').active
            n = 0
            while not sheet[f'A{n + 3}'].value in [None, ''] or str(sheet[f'A{n + 3}'].value).replace(' ', '') == '':
                n += 1  # n is the number of rows with something in it
            rows = []
            for i in range(n):  # loop through all the rows with something in it
                row = []
                # collecting data into 'row'
                for letter in ascii_uppercase[0:9]:
                    value = sheet[f'{letter}{i + 3}'].value
                    try:
                        if letter != 'A':
                            value = float(value)
                    except:
                        if value not in [None, '-', 'N/A']:
                            print(f'{f}: {f"{letter}{i + 3}"} = {value}')
                        value = 0
                    row.append(value)
                # add the row to the sum for this sheet
                rows.append({row[0]: row[1:9]})
            # add this counter to its respective person\
            copy = self.alldata
            name = sheet['B1'].value
            doesitexistalready = False
            for position, personobj in enumerate(copy):
                if personobj['name'] == name:
                    doesitexistalready = True
                    break
            if not doesitexistalready:
                self.alldata.append({'name': name, 'data': rows})
            else:
                for data in rows:
                    self.alldata[position]['data'].append(data)

        pickle.dump(self.alldata, open('data.pickle', 'wb'))

    def output(self, start, end):
        wb = load_workbook('./output_format.xlsx')
        self.load_data()
        sheet = wb.active
        sheet['A1'] = f'{start.strftime("%d/%m/%Y")} - {end.strftime("%d/%m/%Y")}'
        for rowno, advisor in enumerate(self.alldata):
            chunk = {}
            for datachunk in advisor['data']:
                date = list(datachunk.keys())[0]
                data = list(datachunk.values())[0]
                if start <= date <= end:
                    for i, v in enumerate(data):
                        try:
                            chunk[i + 1] += v
                        except:
                            chunk[i + 1] = v
            sheet[f'A{rowno + 3}'] = advisor['name']
            if len(chunk) == 0:
                chunk = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, }
            for k, v in chunk.items():
                if k not in [7, 8]:
                    sheet[f'{ascii_uppercase[k]}{rowno + 3}'] = int(v)
                else:
                    sheet[f'{ascii_uppercase[k]}{rowno + 3}'] = v
        wb.save('./output.xlsx')


c = Compiler()


# c.compile1()
# c.output(datetime(2023, 1, 1, 0, 0), datetime(2023, 3, 31, 0, 0))

class UI:
    def __init__(self, master=None):
        # build ui
        toplevel1 = tk.Tk() if master is None else tk.Toplevel(master)
        toplevel1.configure(height=200, width=200)
        frame2 = tk.Frame(toplevel1)
        frame2.configure(height=720, width=640)
        self.Opendir = tk.Button(frame2)
        self.Opendir.configure(
            background="#fac898",
            cursor="arrow",
            font="TkDefaultFont",
            height=5,
            justify="left",
            overrelief="flat",
            text='Open directory',
            width=100)
        self.Opendir.grid(column=0, row=0)
        self.Opendir.configure(command=self.opendir)
        self.Weekly = tk.Button(frame2)
        self.Weekly.configure(
            background="#c1e1c1",
            cursor="arrow",
            font="TkDefaultFont",
            height=5,
            justify="left",
            text='Load weekly files',
            width=100)
        self.Weekly.grid(column=0, row=1)
        self.Weekly.configure(command=self.weekly)
        self.Backup = tk.Button(frame2)
        self.Backup.configure(
            background="#aec6cf",
            cursor="arrow",
            font="TkDefaultFont",
            height=5,
            justify="left",
            overrelief="flat",
            text='Backup',
            width=100)
        self.Backup.grid(column=0, row=2)
        self.Backup.configure(command=self.backup)
        self.Lbackup = tk.Button(frame2)
        self.Lbackup.configure(
            background="#fdfd96",
            cursor="arrow",
            font="TkDefaultFont",
            height=5,
            justify="left",
            overrelief="flat",
            text='Load backup',
            width=100)
        self.Lbackup.grid(column=0, row=3)
        self.Lbackup.configure(command=self.lbackup)
        self.Clear = tk.Button(frame2)
        self.Clear.configure(
            background="#ff6961",
            cursor="arrow",
            font="TkDefaultFont",
            height=5,
            justify="left",
            overrelief="flat",
            text='Clear',
            width=100)
        self.Clear.grid(column=0, row=4)
        self.Clear.configure(command=self.clear)
        frame2.pack(side="left")
        frame3 = tk.Frame(toplevel1)
        frame3.configure(height=720, width=640)
        frame4 = tk.Frame(frame3)
        frame4.configure(height=200, width=200)
        frame5 = tk.Frame(frame4)
        frame5.configure(height=200, width=200)
        label1 = tk.Label(frame5)
        label1.configure(height=5, text='Start')
        label1.pack(side="top")
        frame7 = tk.Frame(frame5)
        frame7.configure(height=200, width=200)
        self.sd = tk.Entry(frame7)
        self.sd.pack(side="left")
        self.sm = tk.Entry(frame7)
        self.sm.pack(side="left")
        self.sy = tk.Entry(frame7)
        self.sy.pack(side="left")
        frame7.pack(side="top")
        frame5.pack(side="left")
        frame6 = tk.Frame(frame4)
        frame6.configure(height=200, padx=3, width=200)
        label2 = tk.Label(frame6)
        label2.configure(height=5, text='End')
        label2.pack(side="top")
        frame8 = tk.Frame(frame6)
        frame8.configure(height=200, width=200)
        self.ed = tk.Entry(frame8)
        self.ed.pack(side="left")
        self.em = tk.Entry(frame8)
        self.em.pack(side="left")
        self.ey = tk.Entry(frame8)
        self.ey.pack(side="left")
        frame8.pack(side="top")
        frame6.pack(side="top")
        frame4.pack(anchor="n", side="top")
        self.Generate = tk.Button(frame3)
        self.Generate.configure(
            background="#c5b1ff",
            height=21,
            text='Generate')
        self.Generate.pack(expand="true", fill="both", side="top")
        self.Generate.configure(command=self.generate)
        frame3.pack(side="top")

        # Main widget
        self.mainwindow = toplevel1

    def run(self):
        self.mainwindow.mainloop()

    def opendir(self):
        os.system(f'explorer {os.getcwd()}')

    def weekly(self):
        self.backup()
        c.compile2()
        for r, d, fs in os.walk('./thisweek'):
            dest = str(datetime.now()).replace(":", "_")
            os.makedirs(f'./oldfiles/{dest}')
            for f in fs:
                shutil.move(f'./thisweek/{f}', f'./oldfiles/{dest}/{f}')

    def backup(self):
        shutil.copy('data.pickle', f'old/{str(datetime.now()).replace(":", "_")}.pickle')

    def lbackup(self):
        file = askopenfilename()
        self.backup()
        try:
            c.load_data(file)
        except:
            pyag.alert('invalid file')
        shutil.copy(file, './data.pickle')

    def clear(self):
        self.backup()
        shutil.copy('blankdata.pickle', 'data.pickle')
        c.load_data()

    def generate(self):
        try:
            s = datetime(int(self.sy.get()), int(self.sm.get()), int(self.sd.get()), 0, 0)
            e = datetime(int(self.ey.get()), int(self.em.get()), int(self.ed.get()), 0, 0)
            c.output(s, e)
            os.system(f'output.xlsx')
        except Exception as ex:
            pyag.alert(f'invalid date or something else went wrong\n\n{ex}')


if __name__ == "__main__":
    app = UI()
    app.run()
